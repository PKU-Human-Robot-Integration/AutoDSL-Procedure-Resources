import json
import os
import openai
import random
import csv
import time

import pandas as pd
import scipy.stats as stats
from data_process.statement import Statement
from algorithm.EM_feature import EM_feature
from algorithm.Compiler import Compiler
from algorithm.Compiler1 import Compiler as Compiler1
from openai import OpenAI
from utils.token_count_decorator import token_count_decorator
from utils.util import load_json, reorg_semantic_data
from utils.util import write_tsv
import numpy as np
from openpyxl import load_workbook

class Utility_Analyse:
    def __init__(self) -> None:
        self.name_mapping = {
            "Molecular Biology & Genetics": "Genetics",
            "Biomedical & Clinical Research": "Medical",
            "Ecology & Environmental Biology": "Ecology",
            "Bioengineering & Technology": "BioEng",
            "Bioinformatics & Computational Biology": "InfoBio"
        }
        with open("data/syntax_gpt_bio2.txt", 'r') as file:
            self.syntax_extraction_prompt_bio = file.read()
        with open("data/c_feature_list.txt", 'r') as file:
            self.c_feature_list = file.read()
        with open("data/python_feature_list.txt", 'r') as file:
            self.python_feature_list = file.read()
        self.datasets = load_json("data/utility_demo.json")
        self.sub_feature_list = str(['for-loop', 'while-loop', 'if-branch', 'allocate-statement', 'deallocate-statement', 'function-procedure-call', 'function-procedure-declaration', 'add-arithmetic-operator', 'minus-arithmetic-operator', 'multi-arithmetic-operator', 'divide-arithmetic-operator', 'and-arithmetic-operator', 'or-arithmetic-operator', 'not-arithmetic-operator', 'assignment-expression', 'raise-statement', 'integer-type-declaration', 'floating-point-type-declaration', 'boolean-type-declaration', 'string-type-declaration', 'vector-type-declaration', 'set-type-declaration', 'temporal-type-declaration', 'reg-type-declaration', 'device-type-declaration', 'container-type-declaration', 'scientific-type-declaration', 'class-type', 'parallel-for', 'event-statement', 'response-statement'])
        self.specification = {
            "interger-type-declaration" : "If an integer is found in the context, we can consider the language feature \"interger-type-declaration\" to be present.",
            "floatingpoint-type-declaration" : "If a floating-point number is found, we can consider the language feature \"floatingpoint-type-declaration\" to be present.",
            "boolean-type-declaration": "If a boolean value (true or false) is found, we can consider the language feature \"boolean-type-declaration\" to be present.",
            "string-type-declaration": "If a string is found, we can consider the language feature \"string-type-declaration\" to be present.",
            "vector-type-declaration": "If a noun in the context is similar to \"vector,\" we can consider the language feature \"vector-type-declaration\" to be present.",
            "dict-type-declaration": "If a noun in the context is similar to \"dictionary,\" we can consider the language feature \"dict-type-declaration\" to be present.",
            "set-type-declaration": "If a noun in the context is similar to \"set,\" we can consider the language feature \"set-type-declaration\" to be present.",
            "temporal-type-declaration": "If a time-related word is found, we can consider the language feature \"temporal-type-declaration\" to be present.",
            "reg-type-declaration": "If a noun belongs to a set of chemicals, we can consider the language feature \"reg-type-declaration\" to be present.",
            "device-type-declaration": "If a device occurs, we can consider the language feature \"device-type-declaration\" to be present.",
            "container-type-declaration": "If a noun belongs to a set of containers, we can consider the language feature \"container-type-declaration\" to be present.",
            "scientific-type-declaration": "If a scientific type (like volumn) occurs, we can consider the language feature \"scientific-type-declaration\" to be present.",
            "for-loop": "If a word 'repeat' is found in context, we can consider the language feature \"for-loop\" to be present.",
            "while-loop": "If a word 'repeat' is found in context, and there is no integer declaration, we can consider the language feature \"while-loop\" to be present.",
            "if-branch": "If a word 'if' is found in context, we can consider the language feature \"if-branch\" to be present.",
            "if-else-branch": "If word 'if' and word 'else' are found in context, we can consider the language feature \"if-else-branch\" to be present.",
            "function-procedure-call": "If a function or procedure call occurs, we can consider the language feature \"function-procedure-call\" to be present.",
            "function-procedure-declaration": "If a function or procedure call occurs, we can consider the language feature \"function-procedure-declaration\" to be present.",
            "break-statement": "If a while-loop and if-branch are both present, we can consider the language feature \"break-statement\" to be present. If a 'repeat' is found in context, and there is no integer declaration, we can consider the language feature \"while-loop\" to be present. If a word 'if' is found in context, we can consider the language feature \"if-branch\" to be present.",
            "allocate-statement": "If a noun belongs to chemicals or containers, we can consider the language feature \"allocate-statement\" to be present.",
            "deallocate-statement": "If a noun belongs to chemicals or containers, we can consider the language feature \"deallocate-statement\" to be present.",
            "add-arithmetic-operator": "If an addition operation occurs, we can consider the language feature \"add-arithmetic-operator\" to be present.",
            "minus-arithmetic-operator": "If a subtraction operation occurs, we can consider the language feature \"minus-arithmetic-operator\" to be present.",
            "multi-arithmetic-operator": "If a multiplication operation occurs, we can consider the language feature \"multi-arithmetic-operator\" to be present.",
            "devid-arithmetic-operator": "If a division operation occurs, we can consider the language feature \"devid-arithmetic-operator\" to be present.",
            "and-arithmetic-operator": "If an AND operation occurs, we can consider the language feature \"and-arithmetic-operator\" to be present.",
            "or-arithmetic-operator": "If an OR operation occurs, we can consider the language feature \"or-arithmetic-operator\" to be present.",
            "not-arithmetic-operator": "If a NOT operation occurs, we can consider the language feature \"not-arithmetic-operator\" to be present.",
            "assignment-expression": "If a word 'equal' occurs, we can consider the language feature \"assignment-expression\" to be present.",
            "raise-statement": "If a noun in the context is similar to \"error,\" we can consider the language feature \"raise-statement\" to be present.",
            "resolve-statement": "If resolving an issue occurs and a noun similar to \"error\" is present, we can consider the language feature \"resolve-statement\" to be present.",
            "class-type": "If a chemicals occurs in many steps, we can consider the language feature \"class-type\" to be present.",
            "event-statement": "If a word 'when' occurs, we can consider the language feature \"event-statement\" to be present.",
            "response-statement": "If a word 'when' occurs, we can consider the language feature \"response statement\" to be present.",
            "parallel-for": "If an action is followed by many numbers, we can consider the language feature \"parallel-for\" to be present."
        }
        self.c_production_rules = {
            'for-loop': 'for (initialization; condition; increment) statement',
            'while-loop': 'while (condition) statement',
            'if-branch': 'if (condition) statement [else statement]',
            'allocate-statement': 'new type',
            'deallocate-statement': 'delete pointer',
            'function-procedure-call': 'functionName(arguments)',
            'function-procedure-declaration': 'returnType functionName(parameterList) { body }',
            'add-arithmetic-operator': 'expression + expression',
            'minus-arithmetic-operator': 'expression - expression',
            'multi-arithmetic-operator': 'expression * expression',
            'devid-arithmetic-operator': 'expression / expression',
            'and-arithmetic-operator': 'expression && expression',
            'or-arithmetic-operator': 'expression || expression',
            'not-arithmetic-operator': '!expression',
            'assignment-expression': 'variable = expression',
            'raise-statement': 'throw exception',
            'interger-type-declaration': 'int variableName;',
            'floatingpoint-type-declaration': 'float variableName; | double variableName;',
            'boolean-type-declaration': 'bool variableName;',
            'string-type-declaration': 'std::string variableName;',
            'vector-type-declaration': 'std::vector<type> variableName;',
            'set-type-declaration': 'std::set<type> variableName;',
            'class-type': 'class ClassName { ... };',
            'event-statement': 'void onEvent(EventType event) { ... }',
            'response-statement': 'In response to an event: if (event.type == expectedType) { ... }',
            'temporal-type-declaration': 'Using std::chrono for time-related types: std::chrono::time_point<std::chrono::system_clock> temporalVar;',
            'reg-type-declaration': 'Hypothetical reagent type (hardware-related)',
            'device-type-declaration': 'For device management, Device deviceVar;',
            'container-type-declaration': 'Generic container (e.g., from the Standard Template Library): std::container<Type> containerVar;',
            'scientific-type-declaration': 'For scientific datatype, assuming use of a specific library: Volume',
            'parallel-for': 'Using OpenMP for parallel loops: #pragma omp parallel for for (int i = 0; i < N; ++i) { ... }'
        }
        self.python_production_rules = {
            "for-loop": "for variable in iterable: statements",
            "while-loop": "while condition: statements",
            "if-branch": "if condition: statements [elif condition: statements] [else: statements]",
            "allocate-statement": "variable = new Type(arguments) # Not Pythonic, C++-like for memory allocation",
            "deallocate-statement": "del variable # In Python, garbage collection is automatic",
            "function-procedure-call": "function_name(arguments)",
            "function-procedure-declaration": "def function_name(parameters): statements",
            "add-arithmetic-operator": "a + b",
            "minus-arithmetic-operator": "a - b",
            "multi-arithmetic-operator": "a * b",
            "devid-arithmetic-operator": "a / b",
            "and-arithmetic-operator": "a and b",
            "or-arithmetic-operator": "a or b",
            "not-arithmetic-operator": "not a",
            "assignment-expression": "variable = expression",
            "raise-statement": "raise ExceptionType('message')",
            "interger-type-declaration": "variable: int = value # Python 3.6+ type hints",
            "floatingpoint-type-declaration": "variable: float = value",
            "boolean-type-declaration": "variable: bool = value",
            "string-type-declaration": "variable: str = value",
            "vector-type-declaration": "variable: list[Type] = [] # Type hint for a list",
            "set-type-declaration": "variable: set[Type] = set()",
            "temporal-type-declaration": "variable: datetime = datetime.now() # Using datetime module",
            "reg-type-declaration": "variable = re.compile(pattern) # Regular expression compilation",
            "device-type-declaration": "Device device",
            "container-type-declaration": "variable: ContainerType[Type] = ContainerType() # Generic container",
            "scientific-type-declaration": "For scientific datatype, assuming use of a specific library: Volume",
            "class-type": "class ClassName: methods and attributes",
            "parallel-for": "with concurrent.futures.ThreadPoolExecutor() as executor: results = list(executor.map(function, iterable))",
            "event-statement": "def on_event(event): handle_event() # Define an event handler function",
            "response-statement": "def respond_to_event(): return response # Function to create and return a response"
        }

        self.syntax_compiler_datasets = {x:sorted([self.datasets[x][y][0] for y in self.datasets[x]], key=len) for x in self.datasets}
        self.len_syntax_compiler_datasets = 31

    def syntax_utility(self):
        log = load_json("data/syntax_log_3.json", create=True)
        
        for subject in self.datasets:
            print(subject)
            if subject not in log:
                log[subject] = {}

            result_feature = []
            result_ours = []
            result_strong_baseline_biocoder = []
            result_strong_baseline_python = []
            result_medium_baseline = []
            result_weak_baseline_3 = []
            result_weak_baseline_4 = []

            dataset = self.datasets[subject]
            for phi in dataset:
                print(phi)
                protocol = dataset[phi][0]
                result_feature.append(phi)

                another_subject = random.choice([x for x in list(self.datasets.keys()) if x != subject])
                yes_example = self.datasets[another_subject][phi][0]
                another_phi = random.choice([x for x in list(dataset.keys()) if x != phi])
                no_example = self.datasets[another_subject][another_phi][0]

                if phi not in log[subject]:
                    log[subject][phi] = {}

                if "ours" not in log[subject][phi]:
                    content = self.syntax_extraction_prompt_bio.replace("^^^^^^", yes_example).replace("------", self.specification[phi]).replace("++++++", "The production rule:")
                    content2 = self.syntax_extraction_prompt_bio.replace("^^^^^^", no_example).replace("------", self.specification[phi]).replace("++++++", "The production rule:")
                    content4 = self.syntax_extraction_prompt_bio.replace("^^^^^^", protocol).replace("------", self.specification[phi]).replace("++++++", "The production rule:")
                    result = ""
                    while "yes" not in result and "no" not in result:
                        result = self.__chatgpt_function(content, "yes", content2, "no", content4).lower()
                    result_ours.append(1 if "yes" in result.lower() else 0)
                    log[subject][phi]["ours"] = result
                else:
                    result = log[subject][phi]["ours"]
                    result_ours.append(1 if "yes" in result.lower() else 0)
                    
                if "strong baseline-biocoder" not in log[subject][phi]:
                    content = self.syntax_extraction_prompt_bio.replace("^^^^^^", yes_example).replace("------", self.c_production_rules[phi]).replace("++++++", "The production rule:")
                    content2 = self.syntax_extraction_prompt_bio.replace("^^^^^^", no_example).replace("------", self.c_production_rules[phi]).replace("++++++", "The production rule:")
                    content4 = self.syntax_extraction_prompt_bio.replace("^^^^^^", protocol).replace("------", self.c_production_rules[phi]).replace("++++++", "The production rule:")
                    result = ""
                    while "yes" not in result and "no" not in result:
                        result = self.__chatgpt_function(content, "yes", content2, "no", content4).lower()
                    result_strong_baseline_biocoder.append(1 if "yes" in result.lower() else 0)
                    log[subject][phi]["strong baseline-biocoder"] = result
                else:
                    result = log[subject][phi]["strong baseline-biocoder"]
                    result_strong_baseline_biocoder.append(1 if "yes" in result.lower() else 0)

                if "strong baseline-python" not in log[subject][phi]:
                    content = self.syntax_extraction_prompt_bio.replace("^^^^^^", yes_example).replace("------", self.python_production_rules[phi]).replace("++++++", "The production rule:")
                    content2 = self.syntax_extraction_prompt_bio.replace("^^^^^^", no_example).replace("------", self.python_production_rules[phi]).replace("++++++", "The production rule:")
                    content4 = self.syntax_extraction_prompt_bio.replace("^^^^^^", protocol).replace("------", self.python_production_rules[phi]).replace("++++++", "The production rule:")
                    result = ""
                    while "yes" not in result and "no" not in result:
                        result = self.__chatgpt_function(content, "yes", content2, "no", content4).lower()
                    result_strong_baseline_python.append(1 if "yes" in result.lower() else 0)
                    log[subject][phi]["strong baseline-python"] = result
                else:
                    result = log[subject][phi]["strong baseline-python"]
                    result_strong_baseline_python.append(1 if "yes" in result.lower() else 0)
                
                if "medium baseline" not in log[subject][phi]:
                    content = self.syntax_extraction_prompt_bio.replace("^^^^^^", yes_example).replace("------", "Please extract the experimental procedure entities before conducting the analysis.").replace("++++++", "Please identify " + phi)
                    content2 = self.syntax_extraction_prompt_bio.replace("^^^^^^", no_example).replace("------", "Please extract the experimental procedure entities before conducting the analysis.").replace("++++++", "Please identify " + phi)
                    content4 = self.syntax_extraction_prompt_bio.replace("^^^^^^", protocol).replace("------", "Please extract the experimental procedure entities before conducting the analysis.").replace("++++++", "Please identify " + phi)
                    result = ""
                    while "yes" not in result and "no" not in result:
                        result = self.__chatgpt_function(content, "yes", content2, "no", content4).lower()
                    result_medium_baseline.append(1 if "yes" in result.lower() else 0)
                    log[subject][phi]["medium baseline"] = result
                else:
                    result = log[subject][phi]["medium baseline"]
                    result_medium_baseline.append(1 if "yes" in result.lower() else 0)

                if "weak baseline-3" not in log[subject][phi]:
                    content = self.syntax_extraction_prompt_bio.replace("^^^^^^", yes_example).replace("------", "").replace("++++++", "Please identify " + phi)
                    content2 = self.syntax_extraction_prompt_bio.replace("^^^^^^", no_example).replace("------", "").replace("++++++", "Please identify " + phi)
                    content4 = self.syntax_extraction_prompt_bio.replace("^^^^^^", protocol).replace("------", "").replace("++++++", "Please identify " + phi)
                    result = ""
                    while "yes" not in result and "no" not in result:
                        result = self.__chatgpt_function(content, "yes", content2, "no", content4).lower()
                    result_weak_baseline_3.append(1 if "yes" in result.lower() else 0)
                    log[subject][phi]["weak baseline-3"] = result
                else:
                    result = log[subject][phi]["weak baseline-3"]
                    result_weak_baseline_3.append(1 if "yes" in result.lower() else 0)

                if "weak baseline-4" not in log[subject][phi]:
                    content = self.syntax_extraction_prompt_bio.replace("^^^^^^", yes_example).replace("------", "").replace("++++++", "Please identify " + phi)
                    content2 = self.syntax_extraction_prompt_bio.replace("^^^^^^", no_example).replace("------", "").replace("++++++", "Please identify " + phi)
                    content4 = self.syntax_extraction_prompt_bio.replace("^^^^^^", protocol).replace("------", "").replace("++++++", "Please identify " + phi)
                    result = ""
                    while "yes" not in result and "no" not in result:
                        result = self.__chatgpt_function(content, "yes", content2, "no", content4, model="gpt-4").lower()
                    result_weak_baseline_4.append(1 if "yes" in result.lower() else 0)
                    log[subject][phi]["weak baseline-4"] = result
                else:
                    result = log[subject][phi]["weak baseline-4"]
                    result_weak_baseline_4.append(1 if "yes" in result.lower() else 0)

                result = pd.DataFrame()
                result["feature"] = result_feature
                result["ours"] = result_ours
                result["strong_baseline_biocoder"] = result_strong_baseline_biocoder
                result["strong_baseline_python"] = result_strong_baseline_python
                result["medium_baseline"] = result_medium_baseline
                result["weak_baseline_3"] = result_weak_baseline_3
                result["weak_baseline_4"] = result_weak_baseline_4
                result.to_csv("data/csv/" + self.name_mapping[subject] + "_syntax_utility.csv", index=False)
                with open("data/syntax_log_3.json", "w") as file:
                    json.dump(log, file)
        
            data = []
            for column in result.columns:
                if column != 'feature':
                    count_0 = (result[column] == 0).sum()
                    count_1 = (result[column] == 1).sum()
                    data.append({'method': column, 'category': "right", 'count': count_1})
                    data.append({'method': column, 'category': "wrong", 'count': count_0})
            result_df = pd.DataFrame(data, columns=['method', 'category', 'count'])
            print(subject)
            print(result_df)
        return

    def semantic_translation(self, compiler: Compiler):
        compile_log = load_json("data/semantic_translation_log.json", create=True)
        for subject in self.syntax_compiler_datasets:
            print(subject)
            if subject not in compile_log:
                compile_log[subject] = {}

            for i, protocol in enumerate(self.syntax_compiler_datasets[subject][:5]):
                print(protocol)
                if str(i) in compile_log[subject]:
                    continue

                result, log_info = compiler.compile(subject, protocol)
                print("dsl")
                for statement in result["dsl"]:
                    print(statement.result())

                print("baseline1")
                for statement in result["baseline1"]:
                    print(statement.result())

                print("baseline2")
                for statement in result["baseline2"]:
                    print(statement.result())

                print("baseline3")
                for statement in result["baseline3"]:
                    print(statement.result())
                
                compile_log[subject][str(i)] = log_info
                with open("data/semantic_translation_log.json", "w") as file:
                    json.dump(compile_log, file)

        compile_log_1 = load_json("data/semantic_translation_log_1.json", create=True)
        for subject in self.syntax_compiler_datasets:
            print(subject)
            if subject not in compile_log_1:
                compile_log_1[subject] = {}
            
            for i, protocol in enumerate(self.syntax_compiler_datasets[subject][:5]):
                print(protocol)
                old_compile_log = compile_log[subject][str(i)]["compile"]
                if str(i) in compile_log_1[subject]:
                    continue 

                result, log_info = compiler.compile_1(subject, protocol, i, old_compile_log)

                print("baseline4")
                for statement in result["baseline4"]:
                    print(statement.result())

                print("baseline5")
                for statement in result["baseline5"]:
                    print(statement.result())
                
                compile_log_1[subject][str(i)] = log_info
                with open("data/semantic_translation_log_1.json", "w") as file:
                    json.dump(compile_log_1, file)

    def semantic_translation_new(self, compiler: Compiler1):
        log_name = "semantic_translation_log_new"
        log_path = "data/" + log_name + ".json"
        log_path1 = "data/" + log_name + "_1.json"
        compile_log = load_json(log_path, create=True)
        first_compile_log = load_json("data/semantic_translation_log.json", create=True)
        for subject in self.syntax_compiler_datasets:
            print(subject)
            if subject not in compile_log:
                compile_log[subject] = {}

            for i, protocol in enumerate(self.syntax_compiler_datasets[subject][:5]):
                print(protocol)
                old_compile_log = first_compile_log[subject][str(i)]["compile"]
                if str(i) in compile_log[subject]:
                    continue

                result, log_info = compiler.compile(subject, protocol, i, old_compile_log)
                print("dsl")
                for statement in result["dsl"]:
                    print(statement.result())

                print("baseline1")
                for statement in result["baseline1"]:
                    print(statement.result())

                print("baseline2")
                for statement in result["baseline2"]:
                    print(statement.result())

                print("baseline3")
                for statement in result["baseline3"]:
                    print(statement.result())

                compile_log[subject][str(i)] = log_info
                with open(log_path, "w") as file:
                    json.dump(compile_log, file)

        compile_log_1 = load_json(log_path1, create=True)
        for subject in self.syntax_compiler_datasets:
            print(subject)
            if subject not in compile_log_1:
                compile_log_1[subject] = {}

            for i, protocol in enumerate(self.syntax_compiler_datasets[subject][:5]):
                print(protocol)
                old_compile_log = compile_log[subject][str(i)]["compile"]
                if str(i) in compile_log_1[subject]:
                    continue

                result, log_info = compiler.compile_1(subject, protocol, i, old_compile_log)

                print("baseline4")
                for statement in result["baseline4"]:
                    print(statement.result())

                print("baseline5")
                for statement in result["baseline5"]:
                    print(statement.result())

                compile_log_1[subject][str(i)] = log_info
                with open(log_path1, "w") as file:
                    json.dump(compile_log_1, file)
    
    def semantic_utility(self, path):
        # self.reorg_semantic_data()
        methods = []
        datasets = {}
        now_dataset = ""
        now_protocol = ""
        with open('data/compile_result.tsv', 'r', encoding='utf-8') as tsvfile:
            reader = csv.reader(tsvfile, delimiter='\t')
            
            for i, row in enumerate(reader):
                if i == 0:
                    methods = row[3:]
                else:
                    if row[0] != "":
                        datasets[row[0]] = {}
                        now_dataset = row[0]
                    
                    if row[1] != "":
                        datasets[now_dataset][row[1]] = []
                        now_protocol = row[1]

                    datasets[now_dataset][now_protocol].append({methods[i]:Statement.from_str(row[i+3]) for i in range(len(methods))})

        for subject in datasets:
            print(subject)
            action_error_num = pd.DataFrame()
            action_error_mathod = []
            action_error_categ = []
            action_error_categ_num = []
            action_list = {method:[] for method in methods if method != "ground truth"}

            params_error1_num = pd.DataFrame()
            params_error1_method = []
            params_error1_method_num = []
            params_error2_num = pd.DataFrame()
            params_error2_method = []
            params_error2_method_num = []
            params_num = []

            for protocol in datasets[subject]:
                for sentense in datasets[subject][protocol]:
                    for method in methods:
                        if method != "ground truth":
                            action_list[method].append(sentense[method].opcode == sentense["ground truth"].opcode)
                            error1, error2 = self.__calc_param_error_num(sentense["ground truth"], sentense[method])
                            params_error1_method.append(method)
                            params_error1_method_num.append(error1)
                            params_error2_method.append(method)
                            params_error2_method_num.append(error2)
                            params_num.append(len(sentense["ours"].slot)+1)
            
            print(subject, np.mean(np.array(params_num)))

            for method in methods:
                if method != "ground truth":
                    action_error_mathod.append(method)
                    action_error_categ.append("right")
                    action_error_categ_num.append(sum(action_list[method]))
                    action_error_mathod.append(method)
                    action_error_categ.append("error")
                    action_error_categ_num.append(len(action_list[method]) - sum(action_list[method]))
            
            action_error_num["mathod"] = action_error_mathod
            action_error_num["category"] = action_error_categ
            action_error_num["count"] = action_error_categ_num
            action_error_num.to_csv("data/csv/"+self.name_mapping[subject]+"_action_error_num.csv", index=False)
            # print(action_error_num)
            params_error1_num["method"] = params_error1_method
            params_error1_num["error"] = params_error1_method_num
            params_error1_num.to_csv("data/csv/"+self.name_mapping[subject]+"_params_error1_num.csv", index=False)
            params_error2_num["method"] = params_error2_method
            params_error2_num["error"] = params_error2_method_num   
            params_error2_num.to_csv("data/csv/"+self.name_mapping[subject]+"_params_error2_num.csv", index=False)      

    # progarm = [sentense1, sentense2, ...]
    # sentense = [opcode, [paramtype, param], [paramtype, param], ... ,output]
    # ground truth progarm = [a1, a2, ...], opcode set opcode1
    # program = [b1, b2, ...], opcode set opcode2
    # Use of undefined action := |opcode2 - opcode1|
    # Required parameters do not exist := (|a|-|DSL(a,b)|) for pair(a,b)
    # Required parameter underspecified := (|b|-|DSL(a,b)|) for pair(a,b)

    def __calc_param_error_num(self, gt:Statement, a:Statement):
        gt_param_list = gt.slot.copy()
        gt_param_list.append(["output", gt.emit])

        a_param_list = a.slot.copy()
        a_param_list.append(["output", a.emit])

        lcs = self.__calc_lcs(gt_param_list, a_param_list)
        return len(gt_param_list) - lcs, len(a_param_list) - lcs
    
    def __calc_lcs(self, gt, a):
        m = len(gt)
        n = len(a)
        dp = [[0] * (n + 1) for _ in range(m + 1)]
        for i in range(1, m + 1):
            for j in range(1, n + 1):
                if self.__same_param(gt[i - 1], a[j - 1]):
                    dp[i][j] = dp[i - 1][j - 1] + 1
                else:
                    dp[i][j] = max(dp[i - 1][j], dp[i][j - 1])
        return dp[m][n]
    
    def __same_param(self, gt, a):
        if (gt[1] == None or a[1] == None):
            return (gt[0] == a[0])
        else:
            return (str(gt[0]).lower() == str(a[0]).lower() and str(gt[1]).lower() == str(a[1]).lower())
    
    @token_count_decorator
    def __chatgpt_function(self, content, content1, content2, content3, content4, model="gpt-3.5-turbo"):
        while True:
            try:
                client = OpenAI(
                    api_key=os.environ.get("OPENAI_API_KEY"),
                )
                chat_completion = client.chat.completions.create(
                    messages=[
                        {"role": "system", "content": "You are an expert in life science and computer science. Now you are prompted with a grammar of programming language defined by production rules, several experiment steps described in natural language, and a construct (which is the left part in the production rule). Your task is to determine whether the natural language description consists of parts that can be parsed using this production rule. If the natural language string can be parsed, please output \"Yes\", otherwise, outpbut \"No\"."},
                        {"role": "user", "content": content}, 
                        {"role": "assistant", "content": content1}, 
                        {"role": "user", "content": content2}, 
                        {"role": "assistant", "content": content3}, 
                        {"role": "user", "content": content4} 
                    ],
                    model=model,
                )
                return chat_completion.choices[0].message.content
            except openai.APIError as error:
                print(error)

    @token_count_decorator
    def __chatgpt_function_1(self, content, model="gpt-3.5-turbo"):
        while True:
            try:
                client = OpenAI(
                    api_key=os.environ.get("OPENAI_API_KEY"),
                )
                chat_completion = client.chat.completions.create(
                    messages=[
                        {"role": "user", "content": content}
                    ],
                    model=model,
                )
                return chat_completion.choices[0].message.content
            except openai.APIError as error:
                print(error)

    def reorg_semantic_data(self):
        datasets = load_json("data/semantic_translation_log_new.json")
        datasets_1 = load_json("data/semantic_translation_log_new_1.json")
        datasets_2 = []
        wb = load_workbook('data/result_for_expert_checked.xlsx')
        tot_table = [["subject", "protocol", "sentense", "ground truth", "ours", "strong baseline-biocoder",
                      "strong baseline-python", "medium baseline", "weak baseline-gpt3.5", "weak baseline-gpt4"]]
        before_subject = None
        before_protocol = None
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            datasets_2.append(row)
        datasets_2 = datasets_2[1:]
        i = 0
        for subject in datasets:
            for protocol in datasets[subject]:
                data = datasets[subject][protocol]["compile"]
                data_1 = datasets_1[subject][protocol]["compile"]
                for sentense, sentense1 in zip(data, data_1):
                    line = []

                    if "operation" not in sentense:
                        i = i + 1
                        continue
                    else:
                        if subject != before_subject:
                            line.append(subject)
                            before_subject = subject
                        else:
                            line.append("")

                        if protocol != before_protocol:
                            line.append(protocol)
                            before_protocol = protocol
                        else:
                            line.append("")

                        line.append(sentense["sentense"])

                        line.append(datasets_2[i][3])
                        i = i + 1

                        statement = Statement()
                        statement.load_from_dict(sentense["result_dsl"])
                        line.append(statement.result())

                        statement.load_from_dict(sentense1["result_baseline5"])
                        line.append(statement.result())

                        statement.load_from_dict(sentense1["result_baseline4"])
                        line.append(statement.result())

                        statement.load_from_dict(sentense["result_baseline1"])
                        line.append(statement.result())

                        statement.load_from_dict(sentense["result_baseline2"])
                        line.append(statement.result())

                        statement.load_from_dict(sentense["result_baseline3"])
                        line.append(statement.result())

                    tot_table.append(line)

        print(tot_table)
        write_tsv(tot_table, 'data/compile_result.tsv')