import random
import os
import json
import re
import csv
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from data_process.statement import Statement

def load_json(path, create=False, content=None):
    '''
        Load JSON data from a file or create a new JSON file if it doesn't exist.

        @Arguments:
            path (str): The file path to load/create the JSON file.
            create (bool): If True, creates a new JSON file if it doesn't exist. Defaults to False.
            content (dict): The content to be written into the JSON file if create is True. Defaults to None.

        @Returns:
            dict: The JSON data loaded from the file if the file exists.
                If create is True, returns the content that was written to the newly created JSON file.

        @Raises:
            Exception: If the path doesn't exist and create is False.

        @Functionality:
            This function loads JSON data from the specified file path if the file exists.
            If the file doesn't exist and create is True, it creates a new JSON file with the given content.
            If the file doesn't exist and create is False, it raises an exception.
    '''

    if os.path.exists(path):
        with open(path, 'r') as file:
            return json.load(file)
    else:
        if create == True:
            dict = {} if content is None else content
            with open(path, 'w') as file:
                json.dump(dict, file, indent=4)
            return dict
        else:
            raise "path not exist!"
        
def load_csv(path):
    if os.path.exists(path):
        return pd.read_csv(path)
    else:
        return pd.DataFrame()

def seed_set(seed):
    random.seed(seed)
    os.environ['PYTHONHASHSEED'] = str(seed)
    np.random.seed(seed)

def match_and_remove_first_occurrence(main_string: str, sub_string: str):
    '''
        Match and remove the first occurrence of a substring from a main string.

        @Arguments:
            main_string (str): The main string from which the substring will be removed.
            sub_string (str): The substring to be removed from the main string.

        @Output:
            index (int): The index of the first occurrence of the substring in the main string. Returns -1 if the substring is not found.
            new_string (str): The main string after removing the first occurrence of the substring.
    '''
    index = main_string.find(sub_string)

    if index != -1:
        new_string = main_string[:index] + main_string[index + len(sub_string):]
        return index, new_string
    else:
        return index, main_string
    
def normalized_sampling(data):
    '''
        Perform normalized sampling on a given data array.

        @Arguments:
            data (array-like): Input data to be sampled from.

        @Returns:
            tuple: A tuple containing:
                - int: The index of the sampled element from the data.
                - array-like: The normalized data array.
    '''
    normalized_data = data / np.sum(data)
    sampled_index = np.random.choice(len(data), size=1, p=normalized_data)
    return sampled_index[0], normalized_data

def print_pattern(pattern):
    for p in pattern:
        for i, entity_extraction in enumerate(p["example_feature"]):
            corpora = p["example"][i]
            print(str(entity_extraction) + " : " + corpora)
        print("        := " + "\033[91m" + str(p["pattern"]) + "\033[0m")

def check_annotated_format(annotated_corpora):
    '''
        Check Annotated Format in Corpora

        @Arguments:
            annotated_corpora (str): The annotated corpora to be checked.

        @Input:
            annotated_corpora: A string representing annotated corpora with a specific format.

        @Output:
            (tuple): A tuple containing:
                - bool: True if the annotated corpora follows the expected format, False otherwise.
                - origin (list): List of origin annotations extracted from the corpora.
                - label (list): List of label annotations extracted from the corpora.

        @Implementation:
            1. Extract origin annotations and label annotations using regular expressions.
            2. Iterate through origin annotations to find their positions in the corpora and remove them.
            3. Check if the number of origin annotations matches the number of label annotations and if their positions are in the expected order.
            4. Return the result along with extracted origin and label annotations.
    '''
    annotated_corpora = annotated_corpora[:]
    origin = re.findall(r'\[([^[\]]*)\]', annotated_corpora)
    label = re.findall(r'\{([^}]*)\}', annotated_corpora)
    origin_indexs, delete_char_num = [], 0
    for ori in origin:
        index, annotated_corpora = match_and_remove_first_occurrence(annotated_corpora, ori)
        origin_indexs.append(index + delete_char_num)
        if index != -1:
            delete_char_num += len(ori)
    for lab in label:
        _, annotated_corpora = match_and_remove_first_occurrence(annotated_corpora, lab)
    pair = re.findall(r'\[\]\{\}', annotated_corpora)
    if not(len(origin) == len(pair) == len(label) and -1 not in origin_indexs and \
            all(x < y for x, y in zip(origin_indexs, origin_indexs[1:]))):
        return False, origin, label
    else:
        return True, origin, label

def get_list(d):
    '''
        Flatten nested dictionary into a list of keys.

        @Arguments:
            d (dict): Input dictionary with nested structure.

        @Output:
            flattened (dict): Flattened dictionary where keys are concatenated from nested keys, separated by dots.

        @Functionality:
            This function takes a nested dictionary as input and returns a flattened dictionary where keys are concatenated from nested keys, separated by dots. If a value in the input dictionary is itself a dictionary, its keys are recursively flattened and appended to the corresponding parent key. This process continues until all nested dictionaries are flattened.
    '''
    flattened = {}
    for key, value in d.items():
        if isinstance(value, dict):
            flattened[key] = list(value.keys())
            flattened.update(get_list(value))
        else:
            flattened[key] = value
    return flattened

def read_tsv(file_path):
    with open(file_path, 'r', newline='', encoding='utf-8') as tsvfile:
        reader = csv.reader(tsvfile, delimiter='\t')
        for row in reader:
            print(row)

def write_tsv(data, file_path):
    with open(file_path, 'w', newline='', encoding='utf-8') as tsvfile:
        writer = csv.writer(tsvfile, delimiter='\t')
        for row in data:
            writer.writerow(row)

def reorg_semantic_data():
    '''
        Reorganize semantic data from JSON and Excel sources.

        @Inputs:
            No direct inputs. It reads data from the following files:
            - "data/semantic_translation_log.json": JSON file containing semantic translation log.
            - "data/semantic_translation_log_1.json": Another JSON file containing semantic translation log.
            - "data/result_for_expert_checked.xlsx": Excel file containing expert-checked results.

        @Outputs:
            Writes reorganized data to a TSV file named "data/compile_result.tsv".

        @Functionality:
            This function reorganizes semantic data obtained from JSON and Excel sources into a structured format.
            It iterates through the data from the JSON files and matches it with corresponding data from the Excel file.
            For each entry in the JSON data, it retrieves relevant information and combines it with corresponding information from the Excel file.
            The combined data is then structured into a table and written to a TSV file.
    '''

    datasets = load_json("data/semantic_translation_log.json")
    datasets_1 = load_json("data/semantic_translation_log_1.json")
    datasets_2 = []
    wb = load_workbook('data/result_for_expert_checked.xlsx')
    tot_table = [["subject", "protocol", "sentense", "ground truth", "ours", "strong baseline-biocoder", "strong baseline-python", "medium baseline", "weak baseline-gpt3.5", "weak baseline-gpt4"]]
    before_subject = None
    before_protocol = None
    sheet = wb.active
    for row in sheet.iter_rows(values_only=True):
        datasets_2.append(row)
    datasets_2 = datasets_2[1:]
    i = 0
    for subject in datasets:
        for protocol in datasets[subject]:
            data = datasets[subject][protocol]["compile"]
            data_1 = datasets_1[subject][protocol]["compile"]
            for sentense, sentense1 in zip(data, data_1):
                line = []
                
                if "operation" not in sentense:
                    i = i + 1
                    continue
                else:
                    if subject != before_subject:
                        line.append(subject)
                        before_subject = subject
                    else:
                        line.append("")
                    
                    if protocol != before_protocol:
                        line.append(protocol)
                        before_protocol = protocol
                    else:
                        line.append("")

                    line.append(sentense["sentense"])

                    line.append(datasets_2[i][3])
                    i = i + 1

                    statement = Statement()
                    statement.load_from_dict(sentense["result_dsl"])
                    statement.opcode = sentense["operation"]
                    line.append(statement.result())

                    statement.load_from_dict(sentense1["result_baseline5"])
                    line.append(statement.result())

                    statement.load_from_dict(sentense1["result_baseline4"])
                    line.append(statement.result())

                    statement.load_from_dict(sentense["result_baseline1"])
                    line.append(statement.result())

                    statement.load_from_dict(sentense["result_baseline2"])
                    line.append(statement.result())

                    statement.load_from_dict(sentense["result_baseline3"])
                    line.append(statement.result())

                tot_table.append(line)

    print(tot_table)
    write_tsv(tot_table, 'data/compile_result.tsv')